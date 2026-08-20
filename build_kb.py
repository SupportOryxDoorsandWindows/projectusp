#!/usr/bin/env python3
"""
Oryx Doors & Windows - USP Knowledge Base builder.

Reads "Copy of Slider USP.xlsx" and produces:
  data/kb.json   - structured product knowledge base used by the web app
  data/img/*.png - every in-cell technical drawing, extracted and named

Run again whenever the source spreadsheet changes:
    python3 build_kb.py "/path/to/Copy of Slider USP.xlsx"

Notes on the source file
------------------------
Drawings live in this workbook in two different ways, and both have to be
read or real drawings go missing silently:
  - Newer "in-cell images" (rich values). openpyxl reports these cells as
    #VALUE! errors, so they're recovered directly from the .xlsx package
    (xl/richData/*) and mapped back to their cell — see extract_cell_images.
  - Older "floating" pictures anchored over a cell (xl/drawings/*.xml) — a
    plain Insert > Picture, not placed "in" the cell. Every Sliding system's
    drainage drawings are stored this way. See extract_floating_images.
An image present in an option column, by either mechanism and anywhere in
the system's row block, means that option IS supported by the system; a
blank or "N/A" means it is not.
"""

import json
import os
import re
import shutil
import sys
import zipfile

HERE = os.path.dirname(os.path.abspath(__file__))
OUT_DATA = os.path.join(HERE, "data")
OUT_IMG = os.path.join(OUT_DATA, "img")
DEFAULT_SRC = os.path.expanduser("~/Downloads/Copy of Slider USP.xlsx")


# --------------------------------------------------------------------------
# 1. Recover the in-cell drawings and map them to sheet + cell
# --------------------------------------------------------------------------
def extract_cell_images(xlsx_path, dest_dir):
    z = zipfile.ZipFile(xlsx_path)

    def read(p):
        return z.read(p).decode("utf-8")

    rvb = [int(m) for m in re.findall(r'xlrd:rvb i="(\d+)"', read("xl/metadata.xml"))]
    rv_blocks = re.findall(r'<rv s="\d+">(.*?)</rv>', read("xl/richData/rdrichvalue.xml"))
    rv = [int(re.search(r"<v>(\d+)</v>", b).group(1)) for b in rv_blocks]
    rels = re.findall(r'<rel r:id="(rId\d+)"', read("xl/richData/richValueRel.xml"))
    relmap = dict(
        re.findall(
            r'Id="(rId\d+)"[^>]*Target="\.\./media/([^"]+)"',
            read("xl/richData/_rels/richValueRel.xml.rels"),
        )
    )

    wbrels = read("xl/_rels/workbook.xml.rels")
    sheets = re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', read("xl/workbook.xml"))

    mapping = {}
    used = set()
    for name, rid in sheets:
        target = re.search(r'Id="%s"[^>]*Target="(worksheets/[^"]+)"' % rid, wbrels).group(1)
        sheet_xml = read("xl/" + target)
        cells = {}
        for cell, vm in re.findall(r'<c r="([A-Z]+\d+)"[^>]*vm="(\d+)"', sheet_xml):
            media = relmap[rels[rv[rvb[int(vm) - 1]]]]
            fname = "%s_%s%s" % (name.lower(), cell, os.path.splitext(media)[1])
            cells[cell] = fname
            if fname not in used:
                with z.open("xl/media/" + media) as src, open(
                    os.path.join(dest_dir, fname), "wb"
                ) as dst:
                    shutil.copyfileobj(src, dst)
                used.add(fname)
        mapping[name] = cells
    z.close()
    return mapping


def extract_floating_images(xlsx_path, dest_dir):
    """Some drawings in this workbook are older-style "floating" pictures
    anchored over a cell (xl/drawings/*.xml) rather than the newer in-cell
    rich-value images extract_cell_images() reads. Both mechanisms are used
    inconsistently across the file — e.g. every Sliding system's drainage
    drawings are floating pictures — so both need reading, or real drawings
    the technical team already provided get silently dropped."""
    z = zipfile.ZipFile(xlsx_path)

    def read(p):
        return z.read(p).decode("utf-8")

    def col_letter(idx):
        s = ""
        idx += 1
        while idx:
            idx, r = divmod(idx - 1, 26)
            s = chr(65 + r) + s
        return s

    wbrels = read("xl/_rels/workbook.xml.rels")
    sheets = re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', read("xl/workbook.xml"))

    mapping = {}
    used = set()
    for name, rid in sheets:
        sheet_target = re.search(r'Id="%s"[^>]*Target="(worksheets/[^"]+)"' % rid, wbrels).group(1)
        sheet_num = re.search(r"sheet(\d+)\.xml", sheet_target).group(1)
        sheet_rels_path = "xl/worksheets/_rels/sheet%s.xml.rels" % sheet_num
        if sheet_rels_path not in z.namelist():
            continue
        drawing_match = re.search(r'Target="\.\./(drawings/drawing\d+\.xml)"', read(sheet_rels_path))
        if not drawing_match:
            continue
        drawing_num = re.search(r"drawing(\d+)\.xml", drawing_match.group(1)).group(1)
        drawing_rels = read("xl/drawings/_rels/drawing%s.xml.rels" % drawing_num)
        relmap = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="\.\./media/([^"]+)"', drawing_rels))

        cells = {}
        anchors = re.findall(
            r"<xdr:from><xdr:col>(\d+)</xdr:col>.*?<xdr:row>(\d+)</xdr:row>.*?</xdr:from>"
            r'.*?r:embed="(rId\d+)"',
            read("xl/drawings/drawing%s.xml" % drawing_num), re.S)
        for col, row, embed_rid in anchors:
            media = relmap.get(embed_rid)
            if not media:
                continue
            cell = "%s%d" % (col_letter(int(col)), int(row) + 1)
            fname = "%s_%s%s" % (name.lower(), cell, os.path.splitext(media)[1])
            cells[cell] = fname
            if fname not in used:
                with z.open("xl/media/" + media) as src, open(
                    os.path.join(dest_dir, fname), "wb"
                ) as dst:
                    shutil.copyfileobj(src, dst)
                used.add(fname)
        mapping[name] = cells
    z.close()
    return mapping


# --------------------------------------------------------------------------
# 2. Block definitions - which rows and columns belong to which system
# --------------------------------------------------------------------------
COL_SLIDING = {
    "I": ("threshold", "Floor Integrated (marble or wood)"),
    "J": ("threshold", "Floor Flushed"),
    "K": ("threshold", "On Top of FFL"),
    "L": ("threshold", "Stepped Floor"),
    "M": ("drainage", "Visible drainage"),
    "N": ("drainage", "Concealed drainage"),
    "O": ("sightline", "Horizontal sightline"),
    "P": ("sightline", "Vertical sightline"),
}
COL_CASEMENT = {
    "E": ("size", "Size limitation diagram"),
    "G": ("configuration", "Configuration"),
    "H": ("configuration", "Configuration"),
    "I": ("configuration", "Configuration"),
    "J": ("configuration", "Configuration"),
    "K": ("configuration", "Configuration"),
    "L": ("configuration", "Configuration"),
    "N": ("sightline", "Horizontal sightline"),
    "O": ("sightline", "Vertical sightline"),
}
COL_FOLDING = {
    "H": ("configuration", "Configuration"),
    "I": ("configuration", "Configuration"),
    "J": ("configuration", "Configuration"),
    "K": ("configuration", "Configuration"),
    "L": ("configuration", "Configuration"),
    "M": ("configuration", "Configuration"),
    "N": ("drainage", "Visible drainage"),
    "O": ("drainage", "Concealed drainage"),
    "P": ("sightline", "Horizontal sightline"),
    "Q": ("sightline", "Vertical sightline"),
}

# id, sheet, display name, first row, last row, column map,
# and where to read each scalar from
BLOCKS = [
    # --- Sliding ---------------------------------------------------------
    dict(id="series-1", sheet="Sliding", family="Sliding", name="Series 1",
         rows=(3, 8), cols=COL_SLIDING, cfg_col="G", track_col="H",
         w="C3", h="D3", sqm="E3", glass="F3", auto="Q3", lock="R3"),
    dict(id="series-2", sheet="Sliding", family="Sliding", name="Series 2",
         rows=(9, 17), cols=COL_SLIDING, cfg_col="G", track_col="H",
         w="C9", h="D9", sqm="E9", glass="F9", auto="Q9", lock="R9"),
    dict(id="series-3", sheet="Sliding", family="Sliding", name="Series 3 (640)",
         rows=(22, 30), cols=COL_SLIDING, cfg_col="G", track_col="H",
         w="C22", h="D22", sqm="E22", glass="F22", auto="Q22", lock="R22"),
    dict(id="hy40", sheet="Sliding", family="Sliding", name="HY40",
         rows=(31, 36), cols=COL_SLIDING, cfg_col="G", track_col="H",
         w="C31", h="D31", sqm="E31", glass="F31", auto="Q31", lock="R31"),
    # S22 (36 mm / 44 mm glass) removed from the tool at the business's
    # request — rows 39-46 on the Sliding sheet are no longer read. The
    # spreadsheet itself is untouched; re-add these two blocks to bring it
    # back.
    # --- Casement --------------------------------------------------------
    dict(id="hinged-window", sheet="Casement", family="Casement", name="Hinged Window",
         rows=(3, 6), cols=COL_CASEMENT, cfg_col=None, track_col=None,
         w="C3", h="D3", sqm=None, glass="F3", auto="P3", lock="Q3"),
    dict(id="tilt-turn", sheet="Casement", family="Casement", name="Tilt & Turn",
         rows=(7, 10), cols=COL_CASEMENT, cfg_col=None, track_col=None,
         w="C7", h="D7", sqm=None, glass="F7", auto="P3", lock="Q3"),
    dict(id="t8400", sheet="Casement", family="Casement", name="T8400",
         rows=(11, 14), cols=COL_CASEMENT, cfg_col=None, track_col=None,
         w="C11", h="D11", sqm=None, glass="F11", auto="P3", lock="Q3"),
    dict(id="fixed-window", sheet="Casement", family="Casement", name="Fixed Window",
         rows=(19, 22), cols=COL_CASEMENT, cfg_col=None, track_col=None, fixed=True,
         w="C19", h="D19", sqm="E19", glass="F19", auto="P19", lock="Q19"),
    dict(id="single-door", sheet="Casement", family="Casement", name="Single Door",
         rows=(23, 30), cols=COL_CASEMENT, cfg_col=None, track_col=None,
         w="C23", h="D23", sqm="E23", glass="F23", auto="P24", lock="Q23"),
    # --- Folding ---------------------------------------------------------
    dict(id="bifold", sheet="Folding", family="Folding", name="Bifold",
         rows=(3, 8), cols=COL_FOLDING, cfg_col=None, track_col=None,
         leaf_col="G", w="C3", h="D3", sqm="E3", glass="F3", auto="R3", lock="S3"),
    dict(id="reinforced-bifold", sheet="Folding", family="Folding", name="Reinforced Bifold",
         rows=(9, 14), cols=COL_FOLDING, cfg_col=None, track_col=None,
         leaf_col="G", w="C9", h="D9", sqm="E9", glass="F9", auto="R9", lock="S9"),
]


# --------------------------------------------------------------------------
# 3. Helpers
# --------------------------------------------------------------------------
def clean(v):
    if v is None:
        return None
    s = str(v).strip().replace("\n", " ")
    s = re.sub(r"\s+", " ", s)
    if s == "" or s.startswith("#VALUE"):
        return None
    return s


def parse_range(text):
    """'216~1000' -> (216, 1000);  '1500' -> (None, 1500)."""
    if text is None:
        return (None, None)
    t = str(text).replace(",", "").strip()
    m = re.match(r"^(\d+(?:\.\d+)?)\s*[~-]\s*(\d+(?:\.\d+)?)$", t)
    if m:
        return (float(m.group(1)), float(m.group(2)))
    m = re.match(r"^(\d+(?:\.\d+)?)$", t)
    if m:
        return (None, float(m.group(1)))
    return (None, None)


def leaves_in_config(cfg):
    """'Fixed + Sliding + Sliding + Fixed' -> 4 leaves, 2 of them operable."""
    parts = [p.strip().lower() for p in cfg.split("+") if p.strip()]
    return len(parts), sum(1 for p in parts if p.startswith("slid"))


def supported(cell_value, has_image):
    """Tri-state option support.

    True  - a drawing (or text) is present, so the option is offered
    False - the cell says N/A or "- - - -", an explicit no
    None  - the cell is blank: not recorded, which is not the same as no
    """
    if has_image:
        return True
    txt = clean(cell_value)
    if txt is None:
        return None
    if txt.upper() in ("N/A", "NA") or set(txt) <= set("- "):
        return False
    return True


# --------------------------------------------------------------------------
# 4. Build
# --------------------------------------------------------------------------
def build(xlsx_path):
    import openpyxl

    os.makedirs(OUT_IMG, exist_ok=True)
    for f in os.listdir(OUT_IMG):
        os.remove(os.path.join(OUT_IMG, f))

    images = extract_cell_images(xlsx_path, OUT_IMG)
    floating = extract_floating_images(xlsx_path, OUT_IMG)
    for sheet, cells in floating.items():
        sheet_images = images.setdefault(sheet, {})
        for cell, fname in cells.items():
            sheet_images.setdefault(cell, fname)
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    systems = []
    for b in BLOCKS:
        ws = wb[b["sheet"]]
        imap = images.get(b["sheet"], {})
        r0, r1 = b["rows"]

        # Merged ranges only store a value in their top-left cell, so a ref
        # pointing anywhere inside a merge must resolve to that anchor.
        # Folding's Automation (R3:R14) and Locking (S3:S14) span both bifolds.
        merged = {}
        for rng in ws.merged_cells.ranges:
            anchor = ws.cell(row=rng.min_row, column=rng.min_col).coordinate
            for row in ws[rng.coord]:
                for c in row:
                    merged[c.coordinate] = anchor

        def cell(ref):
            if not ref:
                return None
            return clean(ws[merged.get(ref, ref)].value)

        def is_anchor(ref):
            """True unless ref is a non-anchor cell of a merged range. A merged
            configuration cell describes one entry for the whole block, so it
            must be read once, not once per row."""
            return merged.get(ref, ref) == ref

        wmin, wmax = parse_range(cell(b["w"]))
        hmin, hmax = parse_range(cell(b["h"]))
        _, sqm = parse_range(cell(b["sqm"])) if b["sqm"] else (None, None)

        # configurations -------------------------------------------------
        configs = []
        any_config = False
        if b.get("cfg_col"):
            for r in range(r0, r1 + 1):
                ref = "%s%d" % (b["cfg_col"], r)
                if not is_anchor(ref):
                    continue
                cfg = cell(ref)
                if not cfg:
                    continue
                tracks = cell("%s%d" % (b["track_col"], r))
                if cfg.lower().startswith("any"):
                    any_config = True
                    configs.append(dict(label="Any configuration", leaves=None,
                                        operable=None, tracks=tracks))
                elif cfg.upper() == "N/A":
                    configs.append(dict(label="Fixed - not operable", leaves=1,
                                        operable=0, tracks=None))
                else:
                    n, op = leaves_in_config(cfg)
                    configs.append(dict(label=cfg, leaves=n, operable=op, tracks=tracks))
        elif b.get("leaf_col"):
            for r in range(r0, r1 + 1):
                ref = "%s%d" % (b["leaf_col"], r)
                if not is_anchor(ref):
                    continue
                v = cell(ref)
                if v and v.isdigit():
                    n = int(v)
                    configs.append(dict(label="%d-leaf fold" % n, leaves=n,
                                        operable=n, tracks=None))
        elif b.get("fixed"):
            configs.append(dict(label="Fixed — not operable", leaves=1, operable=0, tracks=None))
        else:
            configs.append(dict(label="Single sash", leaves=1, operable=1, tracks=None))

        # tracks available across the block -------------------------------
        tracks = sorted({t for c in configs for t in re.findall(r"\d+", c.get("tracks") or "")},
                        key=int)

        # option columns ---------------------------------------------------
        opts = {"threshold": {}, "drainage": {}, "sightline": {}}
        drawings = []
        for col, (kind, label) in b["cols"].items():
            # A drawing can be anchored anywhere in the block's row range, not
            # only the first row — floating pictures in particular rarely
            # land exactly on r0 — so "is this option offered" has to check
            # the whole block, not just its first row.
            if kind in opts:
                block_has_img = any("%s%d" % (col, r) in imap for r in range(r0, r1 + 1))
                opts[kind][label] = supported(ws["%s%d" % (col, r0)].value, block_has_img)
            for r in range(r0, r1 + 1):
                ref = "%s%d" % (col, r)
                if ref in imap:
                    drawings.append(dict(kind=kind, label=label, file=imap[ref], cell=ref))

        systems.append(dict(
            id=b["id"], name=b["name"], family=b["family"],
            sash_w_min=wmin, sash_w_max=wmax,
            sash_h_min=hmin, sash_h_max=hmax,
            sash_sqm_max=sqm,
            glass=cell(b["glass"]),
            automation=cell(b["auto"]),
            locking=cell(b["lock"]),
            configs=configs, any_config=any_config, tracks=tracks,
            thresholds=opts["threshold"], drainage=opts["drainage"],
            sightlines=opts["sightline"], drawings=drawings,
        ))

    kb = dict(
        source=os.path.basename(xlsx_path),
        systems=systems,
        engineering=ENGINEERING,
        glossary=GLOSSARY,
    )
    with open(os.path.join(OUT_DATA, "kb.json"), "w") as f:
        json.dump(kb, f, indent=1)
    # kb.js lets index.html open straight from the file system, with no server
    with open(os.path.join(OUT_DATA, "kb.js"), "w") as f:
        f.write("window.ORYX_KB = ")
        json.dump(kb, f, indent=1)
        f.write(";\n")

    print("Systems: %d" % len(systems))
    print("Drawings: %d" % len(os.listdir(OUT_IMG)))
    for s in systems:
        print("  %-22s %s x %s mm, %d config(s), %d drawing(s)"
              % (s["name"], s["sash_w_max"], s["sash_h_max"],
                 len(s["configs"]), len(s["drawings"])))


# --------------------------------------------------------------------------
# 5. Engineering notes confirmed by the technical team (not in the sheet)
# --------------------------------------------------------------------------
# Source: technical review conversation, Series 1. Add further systems here as
# the technical team confirms them. Anything not recorded here is reported by
# the bot as "not in the approved knowledge base".
ENGINEERING = {
    "series-1": {
        "Frame depth": "133 mm",
        "Bottom track width": "133 mm",
        "Visible side jamb": "42 mm (overall side frame section 83 mm: 42 mm visible + 41 mm internal return)",
        "Visible interlock": "26 mm between two sliding panels",
        "Sash bottom sightline": "47 mm",
        "Floor Flushed - what you see": "Bottom track hidden. Sash 47 mm visible.",
        "On Top of FFL - what you see": "Bottom track 42 mm + sash 47 mm = 89 mm visible.",
        "Stepped Floor - what you see": "Inside reads as Floor Flushed. Outside reads as On Top of FFL: 42 mm track + 47 mm sash = 89 mm visible externally.",
        "Floor Integrated": "Not available on Series 1.",
        "Hardware": "Latch handle with latch lock mechanism. Pop-out lock is also available.",
        "Locking options": "Latch lock and pop-out lock available.",
        "Drainage options": "Visible, concealed and gutter drainage available.",
        "Interlock and fabrication": "The 26 mm interlock is a visible sightline. Do not treat it automatically as a fabrication deduction.",
    },
    "series-2": {
        "Locking options": "Latch lock and pop-out lock available.",
        "Drainage options": "Visible, concealed and gutter drainage available.",
        "Track depth": "2-track: 167 mm. 3-track: 240 mm. 4-track: 329 mm.",
        "R62 visible sash": "45 mm visible sash sightline with flush track. 80 mm visible sash with top track. If above FFL: 102 mm.",
    },
    "series-3": {
        "Glass stock": "Current stock available for 36 mm glass. 40 mm glass requires new glazing bead stock, ordered depending on the design.",
        "Drainage options": "Visible and concealed drainage available.",
    },
    "hinged-window": {
        "Glass thickness — pending spreadsheet correction": "Confirmed at 30 mm by the technical team; the spec above still reads 28 mm until the source spreadsheet's Glass Thickness column is corrected and rebuilt.",
    },
    "tilt-turn": {
        "Glass thickness — pending spreadsheet correction": "Confirmed at 30 mm by the technical team; the spec above still reads 28 mm until the source spreadsheet's Glass Thickness column is corrected and rebuilt.",
    },
    "fixed-window": {
        "Drainage": "Visible drainage cap available.",
    },
    "single-door": {
        "Maximum sash size — pending spreadsheet correction": "Confirmed at 1,200 × 3,500 mm by the technical team; the spec above still reads 1,200 × 3,000 mm until the source spreadsheet's height limit is corrected and rebuilt.",
        "Threshold options": "Mobility threshold and powder-coated flush threshold available.",
    },
}

GLOSSARY = {
    "Visible drainage": "Water is taken from the bottom-track chamber to an outlet visible on the external face. The outlet can be inspected and accessed for maintenance.",
    "Concealed drainage": "Water is taken through a hidden path so the outlet is not visible externally. Cleaner external appearance.",
    "Floor Flushed": "The bottom track sits level with the finished floor, so no track is visible from inside.",
    "On Top of FFL": "The bottom track sits on top of the finished floor level and is visible.",
    "Stepped Floor": "Flush on the inside, raised track on the outside.",
    "Floor Integrated": "The bottom track is infilled with the floor finish (marble or wood) so the track reads as part of the floor.",
    "Sightline": "The visible width of a profile as seen in elevation. Also called visible dimensions.",
    "Interlock": "The central meeting section between two sliding panels.",
    "Sash": "The moving (or fixed) glazed panel held in the frame.",
}


if __name__ == "__main__":
    src = sys.argv[1] if len(sys.argv) > 1 else DEFAULT_SRC
    if not os.path.exists(src):
        sys.exit("Source spreadsheet not found: %s" % src)
    build(src)
